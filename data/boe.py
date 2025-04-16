#!/usr/bin/env python3
#
# Copyright (c) 2024 LateGenXer
#
# SPDX-License-Identifier: AGPL-3.0-or-later
#


import datetime
import logging
import math
import os
import posixpath
import sys
import zipfile

import openpyxl

import numpy as np
import pandas as pd

from download import download


def read(sh: openpyxl.worksheet.worksheet.Worksheet, data:dict[float, float]) -> datetime.date:
    # Find the last row
    for row in range(6, sh.max_row + 1):
        if sh.cell(row + 1, 1).value is None:
            break

    datetime_ = sh.cell(row, 1).value
    assert isinstance(datetime_, datetime.datetime)
    date = datetime_.date()

    years_row = 4

    assert sh.cell(years_row, 1).value == 'years:'

    for col in range(2, sh.max_column):
        years = sh.cell(years_row, col).value
        if years is None:
            break
        assert isinstance(years, (float, int))

        months = round(years*12)
        assert math.isclose(years*12, months, rel_tol=1e-5)
        years = months / 12.0

        value = sh.cell(row, col).value
        try:
            rate = float(value)  # type: ignore[arg-type]
        except (ValueError, TypeError):
            rate = math.nan

        data[years] = rate

    return date


_data_dir = os.path.dirname(__file__)
_filename = os.path.join(_data_dir, 'boe-yield-curves.csv')


_measures = {
    'Nominal':   'GLC Nominal',
    'Real':      'GLC Real',
    'Inflation': 'GLC Inflation',
    'OIS':       'OIS',
}


def load() -> None:
    url = 'https://www.bankofengland.co.uk/-/media/boe/files/statistics/yield-curves/latest-yield-curve-data.zip'
    filename = os.path.join(_data_dir, posixpath.basename(url))
    download(url, filename, content_type='application/x-zip-compressed', ttl=6*3600)
    archive = zipfile.ZipFile(filename)

    dfs = []

    for measure, name in _measures.items():
        filename = f'{name} daily data current month.xlsx'
        stream = archive.open(filename, 'r')
        wb = openpyxl.load_workbook(stream, read_only=True, data_only=True)

        data: dict[float, float] = {}
        for sheet_name in ['3. spot, short end', '4. spot curve']:
            sh = wb[sheet_name]

            # TODO: Include date
            date = read(sh, data)

        df = pd.DataFrame(data.items(), columns=['Years', f'{measure}_Spot'])
        df.set_index('Years', inplace=True)

        assert df.index.is_monotonic_increasing
        assert df.index.is_unique

        dfs.append(df)

    assert df.index.is_unique
    df = pd.concat(dfs, axis=1)

    assert df.index.is_monotonic_increasing
    assert df.index.is_unique

    df.interpolate(method='cubicspline', axis=0, limit_direction='both', inplace=True)

    # Add reference date
    ref_date = pd.DataFrame([(0.0, date)], columns=['Years', 'Date'])
    ref_date.set_index('Years', inplace=True)
    dfs.append(ref_date)

    assert df.index.is_unique
    df = pd.concat(dfs, axis=1)

    df.sort_index(inplace=True)
    assert df.index.is_monotonic_increasing

    df.to_csv(_filename, float_format='{:.6f}'.format)

    if __name__ == '__main__' and 'plot' in sys.argv[1:]:
        import matplotlib.pyplot as plt
        df.plot(xlim=(0.0, None), grid=True)  # type: ignore[arg-type]
        plt.show()


class Curve:

    def __init__(self, xp:np.ndarray, yp:np.ndarray):
        assert np.all(np.diff(xp) > 0.0)
        assert not np.any(np.isnan(xp))
        assert not np.any(np.isnan(yp))
        self.xp = xp
        self.yp = yp

    def __call__(self, x:float) -> float:
        return float(np.interp(x, self.xp, self.yp))


def YieldCurve(measure:str) -> Curve:
    assert measure in _measures
    download('https://lategenxer.github.io/finance/boe-yield-curves.csv', _filename)
    column = f'{measure}_Spot'
    df = pd.read_csv(_filename, header=0, index_col=0, usecols=['Years', column])
    series = df[column]
    series.dropna(inplace=True)
    xp = series.index.to_numpy()
    yp = series.to_numpy() / 100.0
    return Curve(xp, yp)


if __name__ == '__main__':
    logging.basicConfig(format='%(asctime)s %(name)s %(levelname)s %(message)s', level=logging.INFO)
    load()
