#
# Copyright (c) 2024 LateGenXer
#
# SPDX-License-Identifier: AGPL-3.0-or-later
#


import os.path
import posixpath
import sys
import typing
import threading

import openpyxl

import numpy as np

from download import download


data_dir = os.path.dirname(__file__)


class Table(typing.NamedTuple):

    min_year: int
    max_year: int
    min_age: int
    max_age: int

    array: np.ndarray

    def mortality(self, year:int, age:int) -> float:
        year = max(year, self.min_year)
        year = min(year, self.max_year)
        assert age >= self.min_age
        if age > self.max_age:
            return 1.0
        return self.array[age - self.min_age, year - self.min_year]

    # https://www.ons.gov.uk/peoplepopulationandcommunity/healthandsocialcare/healthandlifeexpectancies/articles/lifeexpectancycalculator/2019-06-07
    def life_expectancy(self, year:int, age:int) -> float:
        yob = age - year
        p = 1.0
        le = 0.0
        for a in range(age, self.max_age, 1):
            year = yob + a
            m = self.mortality(year, age)
            p *= 1.0 - m
            le += p
        return le


def row_values(row:tuple[openpyxl.cell.cell.Cell|openpyxl.cell.cell.MergedCell,...]) -> list:
    return [field.value for field in row]


# Save a NPY atomically
def save_npy(dst:str, array:np.ndarray) -> None:
    dirname, basename = os.path.split(dst)
    tid = threading.get_native_id()
    tmp = os.path.join(dirname, f'.{basename}.{tid}')
    np.save(open(tmp, 'wb'), array)
    os.replace(tmp, dst)


# https://www.ons.gov.uk/peoplepopulationandcommunity/birthsdeathsandmarriages/lifeexpectancies/bulletins/pastandprojecteddatafromtheperiodandcohortlifetables/2020baseduk1981to2070
# https://www.ons.gov.uk/peoplepopulationandcommunity/birthsdeathsandmarriages/lifeexpectancies/datasets/mortalityratesqxprincipalprojectionenglandandwales
# https://www.ons.gov.uk/peoplepopulationandcommunity/birthsdeathsandmarriages/lifeexpectancies/datasets/mortalityratesqxprincipalprojectionunitedkingdom
# TODO: Use or derive unisex tables?
def get_ons_table(basis:str, gender:str) -> Table:
    assert basis in ('period', 'cohort')
    assert gender in ('male', 'female')

    # XXX: Use UK mortality tables?
    url = 'https://www.ons.gov.uk/file?uri=/peoplepopulationandcommunity/birthsdeathsandmarriages/lifeexpectancies/datasets/mortalityratesqxprincipalprojectionenglandandwales/2020based/ewppp20qx.xlsx'
    filename = os.path.join(data_dir, posixpath.basename(url))

    min_year = 1981
    max_year = 2070
    min_age = 0
    max_age = 100
    num_age = max_age - min_age + 1

    wb = None
    npy = os.path.join(data_dir, f'mortality_{basis}_{gender}.npy')
    try:
        if "PYTEST_CURRENT_TEST" in os.environ:
            raise FileNotFoundError
        stream = open(npy, 'rb')
    except FileNotFoundError:
        if wb is None:
            download(url, filename, ttl=sys.maxsize, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)

        sh = wb[f'{gender}s {basis} qx']

        header_cells, = sh.iter_rows(min_row=5, max_row=5)
        header_values = row_values(header_cells)
        assert header_values[0] == 'age'
        assert header_values[1:] == [str(year) for year in range(min_year, max_year + 1)]

        data: list[list] = []
        for row in sh.iter_rows(min_row=6, max_row=6 + num_age - 1):
            assert len(row) == len(header_cells)
            values = row_values(row)
            assert values[0] == min_age + len(data)
            data.append(values[1:])

        assert len(data) == num_age

        array = np.array(data, dtype=np.float32)
        array /= 100000.0

        save_npy(npy, array)
    else:
        array = np.load(stream)

    table = Table(min_year=min_year, max_year=max_year, min_age=min_age, max_age=max_age, array=array)

    return table


# https://www.actuaries.org.uk/learn-and-develop/continuous-mortality-investigation/other-cmi-outputs/unisex-rates-0
def get_cmi_table() -> Table:
    url = "https://www.actuaries.org.uk/system/files/field/document/Unisex%20mortality%20rates%20for%202025-2026%20illustrations%20v01%202024-11-04_0.xlsx"
    filename = os.path.join(data_dir, posixpath.basename(url))
    basis = 'cohort'
    gender = 'unisex'

    min_year = 2024
    max_year = 2124
    num_year = max_year - min_year + 1
    min_age = 20
    max_age = 120
    num_age = max_age - min_age + 1

    npy = os.path.join(data_dir, f'mortality_{basis}_{gender}.npy')
    try:
        if "PYTEST_CURRENT_TEST" in os.environ:
            raise FileNotFoundError
        stream = open(npy, 'rb')
    except FileNotFoundError:
        download(url, filename, ttl=sys.maxsize, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        wb = openpyxl.load_workbook(filename, read_only=True, data_only=True)
        sh = wb['2024-25']

        header_cells, = sh.iter_rows(min_row=4, max_row=4)
        header_values = row_values(header_cells)
        assert header_values[0] == '[x]'
        assert header_values[1 : num_year + 1] == [year for year in range(min_year, max_year + 1)]

        data: list[list[float]] = []
        for row in sh.iter_rows(min_row=5, max_row=5 + num_age - 1):
            assert len(row) == len(header_cells)
            values = row_values(row)
            assert values[0] == min_age + len(data)
            data.append(values[1 : num_year])

        assert len(data) == num_age

        array = np.array(data, dtype=np.float32)
        save_npy(npy, array)
    else:
        array = np.load(stream)

    table = Table(min_year=min_year, max_year=max_year, min_age=min_age, max_age=max_age, array=array)

    return table
